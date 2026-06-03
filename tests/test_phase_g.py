"""
Phase G — Excel dashboard export tests.

Verifies:
  - all 6 tabs present
  - summary contains key metrics including control state
  - open positions tab reflects portfolio state
  - missing prices render as 'n/a' without crashing
  - lock-safe fallback: when the target path can't be opened for write,
    a timestamped sibling file is created and a DashboardResult flags it
  - trades + monitor log tail propagate into their tabs
"""

from datetime import datetime
from pathlib import Path
import pytest
from openpyxl import load_workbook

from bot import control
from bot.config import RiskConfig
from bot.dashboard import (
    DashboardResult, export_dashboard, read_monitor_log_tail,
)
from bot.portfolio import Portfolio, Position


# ── Fixtures ──────────────────────────────────────────────────────────────────
@pytest.fixture
def tmp_control(tmp_path):
    p = tmp_path / "ctrl.json"
    control.set_control_path(p)
    yield p
    control.set_control_path(control.DEFAULT_CONTROL_PATH)


@pytest.fixture
def risk_cfg():
    return RiskConfig()


@pytest.fixture
def fresh_pf():
    return Portfolio.fresh(starting_balance=50_000.0)


def _pos(symbol="ADANIENT", qty=10, entry=2500.0):
    return Position(
        symbol=symbol, qty=qty, entry_price=entry,
        stop=entry * 0.99, target=entry * 1.02,
        charges_buy=15.0, entry_time=datetime.now(),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Tab structure
# ═══════════════════════════════════════════════════════════════════════════════
def test_export_creates_all_six_tabs(tmp_path, tmp_control, fresh_pf, risk_cfg):
    """Phase I added Brain_Learning + Logs; total tabs is now seven."""
    out = tmp_path / "dash.xlsx"
    result = export_dashboard(
        out, portfolio=fresh_pf, risk_cfg=risk_cfg,
        day_start_equity=50_000.0, prices={},
    )
    assert not result.used_fallback
    assert result.written_path == out
    assert out.exists()
    wb = load_workbook(out, read_only=True)
    assert set(wb.sheetnames) == {
        "Summary", "Open Positions", "Trades",
        "Signals_Research", "Brain_Learning", "Risk", "Logs",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Summary contents
# ═══════════════════════════════════════════════════════════════════════════════
def test_summary_includes_core_metrics(tmp_path, tmp_control, fresh_pf, risk_cfg):
    out = tmp_path / "dash.xlsx"
    export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                     day_start_equity=50_000.0, prices={})
    wb = load_workbook(out, read_only=True)
    ws = wb["Summary"]
    keys = {row[0] for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)
            if row and row[0]}
    expected_subset = {
        "Cash", "Equity (live)", "Day-start equity", "Realized P&L",
        "Open / unrealized P&L", "Total trades", "Win rate (%)",
        "Daily loss (%)", "Drawdown (%)", "Peak equity",
        "Open positions", "Control: killed", "Control: paused",
    }
    assert expected_subset.issubset(keys), keys


def test_summary_reflects_control_state(tmp_path, tmp_control, fresh_pf, risk_cfg):
    control.kill("test reason")
    out = tmp_path / "dash.xlsx"
    export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                     day_start_equity=50_000.0, prices={})
    wb = load_workbook(out, read_only=True)
    ws = wb["Summary"]
    rows = {row[0]: row[1] for row in ws.iter_rows(min_row=2, max_col=2,
                                                    values_only=True) if row[0]}
    assert rows["Control: killed"] is True
    assert rows["Control: kill_reason"] == "test reason"


# ═══════════════════════════════════════════════════════════════════════════════
# Open positions tab
# ═══════════════════════════════════════════════════════════════════════════════
def test_open_positions_with_live_price(tmp_path, tmp_control, fresh_pf, risk_cfg):
    fresh_pf.add_position(_pos("ADANIENT", qty=10, entry=2500.0), cost=25_015)
    out = tmp_path / "dash.xlsx"
    export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                     day_start_equity=50_000.0,
                     prices={"ADANIENT": 2550.0})
    wb = load_workbook(out, read_only=True)
    ws = wb["Open Positions"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) >= 1
    first = rows[0]
    assert first[0] == "ADANIENT"
    assert first[1] == 10
    assert first[2] == 2500.0
    assert first[3] == 2550.0
    assert first[6] == pytest.approx(500.0)   # unrealized P&L


def test_open_positions_missing_price_marked_na(tmp_path, tmp_control,
                                                  fresh_pf, risk_cfg):
    fresh_pf.add_position(_pos("ADANIENT"), cost=25_015)
    out = tmp_path / "dash.xlsx"
    export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                     day_start_equity=50_000.0, prices={})   # no price
    wb = load_workbook(out, read_only=True)
    ws = wb["Open Positions"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    first = rows[0]
    assert first[3] == "n/a"
    assert first[6] == "n/a"


# ═══════════════════════════════════════════════════════════════════════════════
# Trades + Monitor Log propagation
# ═══════════════════════════════════════════════════════════════════════════════
def test_trades_propagate(tmp_path, tmp_control, fresh_pf, risk_cfg):
    out = tmp_path / "dash.xlsx"
    sample_trades = [
        {"ts": "2026-05-25 10:00:00", "symbol": "ADANIENT", "action": "BUY",
         "price": 2500.0, "qty": 10, "value": 25000.0, "charges": 15.0,
         "stop": 2475.0, "target": 2550.0, "pnl": None, "balance_after": 24985.0,
         "exit_reason": None, "slippage": 0.0},
        {"ts": "2026-05-25 11:00:00", "symbol": "ADANIENT", "action": "SELL",
         "price": 2550.0, "qty": 10, "value": 25500.0, "charges": 18.0,
         "stop": 2475.0, "target": 2550.0, "pnl": 467.0, "balance_after": 50467.0,
         "exit_reason": "TARGET", "slippage": 1.25},
    ]
    export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                     day_start_equity=50_000.0, trades=sample_trades, prices={})
    wb = load_workbook(out, read_only=True)
    ws = wb["Trades"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2
    assert data_rows[0][1] == "ADANIENT"
    assert data_rows[1][2] == "SELL"
    assert data_rows[1][12] == "TARGET"   # exit_reason column


def test_monitor_log_tail_propagates(tmp_path, tmp_control, fresh_pf, risk_cfg):
    out = tmp_path / "dash.xlsx"
    tail = [
        "[2026-05-25 10:00:00] monitor-once start\n",
        "  monitor-once: no open positions\n",
        "[2026-05-25 10:00:00] monitor-once exit code: 0\n",
    ]
    export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                     day_start_equity=50_000.0, prices={},
                     monitor_log_tail=tail)
    wb = load_workbook(out, read_only=True)
    ws = wb["Logs"]
    rows = [r[0] for r in ws.iter_rows(min_row=2, max_col=1, values_only=True)]
    joined = "\n".join(r for r in rows if r)
    assert "monitor-once start" in joined
    assert "exit code: 0" in joined


# ═══════════════════════════════════════════════════════════════════════════════
# Lock-safe fallback
# ═══════════════════════════════════════════════════════════════════════════════
def test_lock_safe_fallback_when_path_locked(tmp_path, tmp_control,
                                              fresh_pf, risk_cfg, monkeypatch):
    """
    Simulate Excel holding the file open: openpyxl's Workbook.save raises
    PermissionError on write. The exporter must catch it and write to a
    timestamped sibling file.
    """
    out = tmp_path / "dash.xlsx"

    import bot.dashboard as dash
    real_save = dash.Workbook.save
    calls = {"n": 0}

    def fake_save(self, target):
        calls["n"] += 1
        if calls["n"] == 1:
            raise PermissionError("simulated: file in use")
        # Fallback call — actually save
        return real_save(self, target)

    monkeypatch.setattr(dash.Workbook, "save", fake_save)

    result = export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                              day_start_equity=50_000.0, prices={})
    assert result.used_fallback
    assert result.written_path != out
    assert result.written_path.exists()
    assert "locked-" in result.written_path.name
    assert "PermissionError" in (result.fallback_reason or "")


def test_export_never_raises_on_unexpected_error(tmp_path, tmp_control,
                                                  fresh_pf, risk_cfg, monkeypatch):
    """Unexpected exceptions get swallowed and returned in DashboardResult."""
    out = tmp_path / "dash.xlsx"
    import bot.dashboard as dash

    def boom(self, target):
        raise RuntimeError("unexpected disk fault")
    monkeypatch.setattr(dash.Workbook, "save", boom)

    result = export_dashboard(out, portfolio=fresh_pf, risk_cfg=risk_cfg,
                              day_start_equity=50_000.0, prices={})
    assert isinstance(result, DashboardResult)
    assert result.used_fallback
    assert "unexpected" in (result.fallback_reason or "").lower()


# ═══════════════════════════════════════════════════════════════════════════════
# read_monitor_log_tail
# ═══════════════════════════════════════════════════════════════════════════════
def test_monitor_log_tail_missing_returns_empty(tmp_path):
    assert read_monitor_log_tail(tmp_path / "nope.log") == []


def test_monitor_log_tail_returns_last_n(tmp_path):
    p = tmp_path / "m.log"
    p.write_text("\n".join(f"line {i}" for i in range(100)) + "\n",
                 encoding="utf-8")
    tail = read_monitor_log_tail(p, n=5)
    assert len(tail) == 5
    assert tail[-1].strip() == "line 99"
