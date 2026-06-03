"""
Excel dashboard for kite-bot.

Writes a 6-tab workbook summarizing portfolio + control + risk + trades +
monitor activity. Designed to be refreshed after every state-changing command.

File-lock safety: Excel holds an exclusive lock when the workbook is open.
If the canonical path is locked, we write a timestamped fallback file
(same directory) and log a warning. We NEVER crash the engine on a write
failure — dashboards are observational, not operational.
"""

from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bot import control
from bot.config import RiskConfig
from bot.logger_config import get_logger
from bot.portfolio import Portfolio

log = get_logger("kite-bot.dashboard")

DEFAULT_DASHBOARD_PATH = (
    Path(__file__).resolve().parents[1] / "control" / "kite-bot-dashboard.xlsx"
)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


@dataclass
class DashboardResult:
    written_path: Path
    used_fallback: bool
    fallback_reason: Optional[str] = None


# ── Public API ───────────────────────────────────────────────────────────────
def export_dashboard(
    path:              Path = DEFAULT_DASHBOARD_PATH,
    *,
    portfolio:         Portfolio,
    risk_cfg:          RiskConfig,
    day_start_equity:  float,
    prices:            Optional[Dict[str, float]] = None,
    trades:            Optional[List[dict]] = None,
    monitor_log_tail:  Optional[List[str]] = None,
    mode_info:         Optional[dict] = None,
    learner_profile:   Optional[dict] = None,
    heartbeat_tail:    Optional[List[str]] = None,
) -> DashboardResult:
    """
    Render the workbook. On file-lock (Excel has it open), write a
    timestamped sibling file and log a warning. Never raises.
    """
    path = Path(path)
    prices = prices or {}
    trades = trades or []
    monitor_log_tail = monitor_log_tail or []

    wb = _build_workbook(
        portfolio=portfolio, risk_cfg=risk_cfg,
        day_start_equity=day_start_equity, prices=prices,
        trades=trades, monitor_log_tail=monitor_log_tail,
        mode_info=mode_info or {},
        learner_profile=learner_profile or {},
        heartbeat_tail=heartbeat_tail or [],
    )

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return DashboardResult(written_path=path, used_fallback=False)
    except PermissionError as e:
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        fallback = path.with_name(f"{path.stem}.locked-{ts}{path.suffix}")
        try:
            wb.save(fallback)
            log.warning(f"dashboard locked at {path} ({e}); wrote fallback {fallback}")
            return DashboardResult(
                written_path=fallback, used_fallback=True,
                fallback_reason=f"PermissionError: {e}",
            )
        except Exception as e2:
            csv_fallback = _write_csv_fallback(
                path.parent, portfolio, day_start_equity, prices, trades
            )
            log.error(f"dashboard fallback also failed ({e2}); wrote CSV {csv_fallback}")
            return DashboardResult(
                written_path=csv_fallback, used_fallback=True,
                fallback_reason=f"xlsx failed: {e}; {e2}",
            )
    except Exception as e:
        log.exception("dashboard export raised; not crashing engine")
        return DashboardResult(
            written_path=path, used_fallback=True,
            fallback_reason=f"unexpected: {e}",
        )


# ── Workbook construction ────────────────────────────────────────────────────
def _build_workbook(
    *, portfolio: Portfolio, risk_cfg: RiskConfig, day_start_equity: float,
    prices: Dict[str, float], trades: List[dict], monitor_log_tail: List[str],
    mode_info: dict, learner_profile: dict, heartbeat_tail: List[str],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    _build_summary(wb, portfolio, risk_cfg, day_start_equity, prices, mode_info)
    _build_open_positions(wb, portfolio, prices)
    _build_trades(wb, trades)
    _build_signals_research(wb)
    _build_brain_learning(wb, learner_profile)
    _build_risk(wb, portfolio, risk_cfg, day_start_equity, prices)
    _build_logs(wb, monitor_log_tail, heartbeat_tail)
    return wb


def _header(ws, row: int, columns: List[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="left")


def _autosize(ws, width_caps: int = 28) -> None:
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_len + 2, width_caps)


def _safe_equity(pf: Portfolio, prices: Dict[str, float]) -> Optional[float]:
    try:
        return pf.equity(prices)
    except Exception:
        return None


def _safe_unrealized(pf: Portfolio, prices: Dict[str, float]) -> Optional[float]:
    try:
        return pf.unrealized_pnl(prices)
    except Exception:
        return None


def _safe_drawdown(pf: Portfolio, prices: Dict[str, float]) -> Optional[float]:
    try:
        return pf.drawdown_pct(prices)
    except Exception:
        return None


def _safe_gross_exposure(pf: Portfolio, prices: Dict[str, float]) -> Optional[float]:
    try:
        return pf.gross_exposure(prices)
    except Exception:
        return None


# ── Summary tab ──────────────────────────────────────────────────────────────
def _build_summary(ws_parent: Workbook, pf: Portfolio, risk_cfg: RiskConfig,
                   day_start_equity: float, prices: Dict[str, float],
                   mode_info: dict) -> None:
    ws = ws_parent.create_sheet("Summary")
    state = control.read_state()
    equity     = _safe_equity(pf, prices)
    unrealized = _safe_unrealized(pf, prices)
    drawdown   = _safe_drawdown(pf, prices)
    daily_loss_pct = None
    if day_start_equity and equity is not None:
        daily_loss_pct = max(0.0, (day_start_equity - equity) / day_start_equity * 100)

    # Running state inferred from control flags
    if state.killed:
        run_state = "STOPPED (kill switch tripped)"
    elif state.paused:
        run_state = "PAUSED (entries blocked, exits still firing)"
    else:
        run_state = "RUNNING"

    rows = [
        ("Generated at",       datetime.now().isoformat(timespec="seconds")),
        ("Running state",      run_state),
        ("Mode",               mode_info.get("mode", "unknown")),
        ("Asset class",        mode_info.get("asset_class", "")),
        ("Quote currency",     mode_info.get("quote_currency", "")),
        ("Market hours 24x7",  mode_info.get("market_hours_24x7", "")),
        ("Last heartbeat",     mode_info.get("last_heartbeat", "n/a")),
        ("TV launch status",   mode_info.get("tv_launch_status", "n/a")),
        ("Unavailable symbols", ", ".join(mode_info.get("unavailable_symbols", []))),
        ("Cash",               round(pf.state.cash, 2)),
        ("Equity (live)",      round(equity, 2) if equity is not None else "n/a (missing prices)"),
        ("Day-start equity",   round(day_start_equity, 2)),
        ("Realized P&L",       round(pf.state.realized_pnl, 2)),
        ("Open / unrealized P&L", round(unrealized, 2) if unrealized is not None else "n/a"),
        ("Total trades",       pf.state.total_trades),
        ("Winning trades",     pf.state.winning_trades),
        ("Win rate (%)",       round(pf.win_rate_pct, 2)),
        ("Daily loss (%)",     round(daily_loss_pct, 2) if daily_loss_pct is not None else "n/a"),
        ("Drawdown (%)",       round(drawdown, 2) if drawdown is not None else "n/a"),
        ("Peak equity",        round(pf.state.peak_equity, 2)),
        ("Open positions",     len(pf.state.positions)),
        ("Control: killed",    state.killed),
        ("Control: kill_reason", state.kill_reason or ""),
        ("Control: paused",    state.paused),
        ("Control: pause_reason", state.pause_reason or ""),
    ]
    _header(ws, 1, ["Metric", "Value"])
    for r_idx, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=k)
        ws.cell(row=r_idx, column=2, value=v)
    _autosize(ws)


# ── Open Positions tab ───────────────────────────────────────────────────────
def _build_open_positions(ws_parent: Workbook, pf: Portfolio,
                           prices: Dict[str, float]) -> None:
    ws = ws_parent.create_sheet("Open Positions")
    cols = ["Symbol", "Qty", "Entry Price", "Current Price", "Stop", "Target",
            "Unrealized P&L", "Exposure (Rs.)", "Entry Time"]
    _header(ws, 1, cols)
    if not pf.state.positions:
        ws.cell(row=2, column=1, value="(no open positions)")
        _autosize(ws)
        return
    row = 2
    for sym, pos in pf.state.positions.items():
        cur = prices.get(sym)
        unreal = (cur - pos.entry_price) * pos.qty if cur is not None else None
        exposure = cur * pos.qty if cur is not None else pos.entry_price * pos.qty
        ws.cell(row=row, column=1, value=sym)
        ws.cell(row=row, column=2, value=pos.qty)
        ws.cell(row=row, column=3, value=pos.entry_price)
        ws.cell(row=row, column=4, value=cur if cur is not None else "n/a")
        ws.cell(row=row, column=5, value=pos.stop)
        ws.cell(row=row, column=6, value=pos.target)
        ws.cell(row=row, column=7,
                value=round(unreal, 2) if unreal is not None else "n/a")
        ws.cell(row=row, column=8, value=round(exposure, 2))
        ws.cell(row=row, column=9, value=pos.entry_time.isoformat(timespec="seconds"))
        row += 1
    _autosize(ws)


# ── Trades tab ───────────────────────────────────────────────────────────────
def _build_trades(ws_parent: Workbook, trades: List[dict]) -> None:
    ws = ws_parent.create_sheet("Trades")
    cols = ["ts", "symbol", "action", "price", "qty", "value", "charges",
            "stop", "target", "pnl", "balance_after",
            "entry_reason", "exit_reason", "slippage"]
    _header(ws, 1, cols)
    for r_idx, t in enumerate(trades, start=2):
        for c_idx, col in enumerate(cols, start=1):
            ws.cell(row=r_idx, column=c_idx, value=t.get(col))
    if not trades:
        ws.cell(row=2, column=1, value="(no trades yet)")
    _autosize(ws)


# ── Risk tab ─────────────────────────────────────────────────────────────────
def _build_risk(ws_parent: Workbook, pf: Portfolio, risk_cfg: RiskConfig,
                day_start_equity: float, prices: Dict[str, float]) -> None:
    ws = ws_parent.create_sheet("Risk")
    equity = _safe_equity(pf, prices)
    gross  = _safe_gross_exposure(pf, prices)
    gross_pct = (gross / equity * 100) if (gross is not None and equity) else None
    daily_loss_pct = None
    if day_start_equity and equity is not None:
        daily_loss_pct = max(0.0, (day_start_equity - equity) / day_start_equity * 100)
    drawdown = _safe_drawdown(pf, prices)

    rows = [
        ("max_open_positions",      risk_cfg.max_open_positions,
            f"current: {len(pf.state.positions)}"),
        ("max_total_exposure_pct",  risk_cfg.max_total_exposure_pct,
            f"current: {round(gross_pct, 2) if gross_pct is not None else 'n/a'}"),
        ("max_symbol_notional_pct", risk_cfg.max_symbol_notional_pct,
            "per-symbol — see Open Positions"),
        ("max_daily_loss_pct",      risk_cfg.max_daily_loss_pct,
            f"current: {round(daily_loss_pct, 2) if daily_loss_pct is not None else 'n/a'}"),
        ("max_drawdown_pct",        risk_cfg.max_drawdown_pct,
            f"current: {round(drawdown, 2) if drawdown is not None else 'n/a'}"),
        ("risk_per_trade_pct",      risk_cfg.risk_per_trade_pct, ""),
    ]
    _header(ws, 1, ["Cap", "Limit", "Status"])
    for r_idx, (cap, limit, status) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=cap)
        ws.cell(row=r_idx, column=2, value=limit)
        ws.cell(row=r_idx, column=3, value=status)
    _autosize(ws)


# ── Brain_Learning tab ───────────────────────────────────────────────────────
def _build_brain_learning(ws_parent: Workbook, profile: dict) -> None:
    """
    Read-only view of the adaptive scoring profile from bot.learner.
    Empty/missing profile renders a friendly placeholder.
    """
    ws = ws_parent.create_sheet("Brain_Learning")
    _header(ws, 1, ["Metric", "Value"])
    if not profile:
        ws.cell(row=2, column=1, value="(no strategy_profile.json yet — "
                                        "learner runs after enough closed trades)")
        _autosize(ws)
        return
    weights = profile.get("weights", {}) or {}
    rows = [
        ("Total closed trades", profile.get("total_closed_trades", 0)),
        ("Sample size sufficient", profile.get("sample_size_sufficient", False)),
        ("Win rate", profile.get("win_rate", 0.0)),
        ("Avg R multiple", profile.get("avg_r_multiple", 0.0)),
        ("Losing streak", profile.get("losing_streak", 0)),
        ("Max drawdown (%)", profile.get("max_drawdown_pct", 0.0)),
        ("Cooled-down symbols", ", ".join(profile.get("cooled_down_symbols", []))),
        ("Last updated", profile.get("last_updated", "")),
        ("Notes", profile.get("notes", "")),
        ("--- weights ---", ""),
    ]
    for k in ("technical", "sentiment", "fundamentals", "liquidity", "risk"):
        rows.append((f"weight: {k}", weights.get(k, 0.0)))
    for r_idx, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=k)
        ws.cell(row=r_idx, column=2, value=v)
    _autosize(ws)


# ── Logs tab (monitor.log + heartbeat.log combined) ──────────────────────────
def _build_logs(ws_parent: Workbook, monitor_tail: List[str],
                heartbeat_tail: List[str]) -> None:
    ws = ws_parent.create_sheet("Logs")
    _header(ws, 1, ["Monitor log (recent)"])
    row = 2
    if not monitor_tail:
        ws.cell(row=row, column=1, value="(no monitor.log entries)")
        row += 1
    else:
        for line in monitor_tail[-50:]:
            ws.cell(row=row, column=1, value=line.rstrip()); row += 1
    row += 1
    ws.cell(row=row, column=1, value="--- Heartbeat log (recent) ---").font = Font(bold=True)
    row += 1
    if not heartbeat_tail:
        ws.cell(row=row, column=1, value="(no heartbeat.log entries)")
    else:
        for line in heartbeat_tail[-50:]:
            ws.cell(row=row, column=1, value=line.rstrip()); row += 1
    ws.column_dimensions["A"].width = 140


# ── Signals/Research tab ─────────────────────────────────────────────────────
def _build_signals_research(ws_parent: Workbook) -> None:
    """
    Two stacked sections in one sheet:
      1. Latest signal candidates (from signal_candidates table)
      2. Today's research snapshots (from research_snapshots table)
    Both are read-only views — never executed/mutated here.
    """
    from datetime import date as _date
    try:
        from bot.scanner import list_recent_candidates
        from bot.research import list_snapshots_for_date
        candidates = list_recent_candidates(limit=50)
        snapshots  = list_snapshots_for_date(_date.today())
    except Exception:
        candidates = []
        snapshots  = []

    ws = ws_parent.create_sheet("Signals_Research")

    # Section 1 — candidates
    cand_cols = ["ts", "symbol", "signal", "total_score",
                 "technical_score", "sentiment_score", "fundamentals_score",
                 "liquidity_score", "risk_score",
                 "entry_blocked", "block_reasons", "rejection_reason"]
    _header(ws, 1, cand_cols)
    if not candidates:
        ws.cell(row=2, column=1, value="(no signal candidates logged yet — run scan-once)")
        next_row = 3
    else:
        for r_idx, c in enumerate(candidates, start=2):
            for col_idx, col in enumerate(cand_cols, start=1):
                ws.cell(row=r_idx, column=col_idx, value=c.get(col))
        next_row = 2 + len(candidates) + 1

    # Spacer + section 2 — research snapshots
    snap_header_row = next_row + 1
    ws.cell(row=snap_header_row - 1, column=1,
            value="--- Research Snapshots (today) ---").font = Font(bold=True)
    snap_cols = ["symbol", "asof", "fundamentals_score",
                 "sentiment_score", "liquidity_score", "computed_at"]
    _header(ws, snap_header_row, snap_cols)
    if not snapshots:
        ws.cell(row=snap_header_row + 1, column=1,
                value="(no cached snapshots — scan-once will populate)")
    else:
        for r_idx, s in enumerate(snapshots, start=snap_header_row + 1):
            for col_idx, col in enumerate(snap_cols, start=1):
                ws.cell(row=r_idx, column=col_idx, value=s.get(col))
    _autosize(ws)


# ── CSV fallback (last resort) ───────────────────────────────────────────────
def _write_csv_fallback(parent: Path, pf: Portfolio, day_start_equity: float,
                        prices: Dict[str, float], trades: List[dict]) -> Path:
    import csv
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = parent / f"kite-bot-dashboard.fallback-{ts}.csv"
    equity = _safe_equity(pf, prices)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["section", "key", "value"])
        w.writerow(["summary", "generated_at", datetime.now().isoformat()])
        w.writerow(["summary", "cash", pf.state.cash])
        w.writerow(["summary", "equity", equity if equity is not None else "n/a"])
        w.writerow(["summary", "day_start_equity", day_start_equity])
        w.writerow(["summary", "realized_pnl", pf.state.realized_pnl])
        w.writerow(["summary", "open_positions", len(pf.state.positions)])
        for sym, pos in pf.state.positions.items():
            w.writerow(["position", sym, f"qty={pos.qty} entry={pos.entry_price} "
                                          f"stop={pos.stop} target={pos.target}"])
        for t in trades[-50:]:
            w.writerow(["trade", t.get("ts"),
                        f"{t.get('action')} {t.get('symbol')} qty={t.get('qty')} "
                        f"price={t.get('price')} pnl={t.get('pnl')}"])
    return out


def read_monitor_log_tail(path: Path, n: int = 50) -> List[str]:
    """Read last n lines of monitor.log if present. Returns [] on failure."""
    try:
        if not path.exists():
            return []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        return lines[-n:]
    except Exception:
        return []
